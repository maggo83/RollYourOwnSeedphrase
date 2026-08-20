#!/usr/bin/env python3
"""Generate localized printouts without changing the English BIP39 word list."""

from __future__ import annotations

import argparse
import html
import json
import os
import shutil
import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
RESOURCES = ROOT / "additional_ressources"
PRINT_SOURCE = ROOT / "print-src"
QUICK_GUIDE_TEMPLATE = PRINT_SOURCE / "quick-guide-template.html"
QUICK_GUIDE_CATALOG = PRINT_SOURCE / "quick-guide-content.json"
QUICK_GUIDE_SLOTS = (
    "eyebrow", "headline", "lede", "notice", "links", "preparation",
    "step-1", "step-2", "step-3", "step-4", "step-5", "footer",
)


def convert_to_pdf(source: Path, destination: Path) -> None:
    with tempfile.TemporaryDirectory(dir=os.environ.get("XDG_RUNTIME_DIR")) as profile:
        subprocess.run(
            [
                "libreoffice",
                "--headless",
                "--nologo",
                "--nofirststartwizard",
                "--norestore",
                f"-env:UserInstallation=file://{profile}",
                "--convert-to",
                "pdf",
                "--outdir",
                str(destination.parent),
                str(source),
            ],
            check=True,
            env={**os.environ, "TMPDIR": os.environ.get("XDG_RUNTIME_DIR", os.environ.get("TMPDIR", "/tmp"))},
        )
    generated = destination.parent / f"{source.stem}.pdf"
    if generated != destination:
        generated.replace(destination)


def render_html_to_pdf(source: Path, destination: Path) -> None:
    browser = next(
        (candidate for name in ("google-chrome", "chromium", "chromium-browser") if (candidate := shutil.which(name))),
        None,
    )
    if browser is None:
        raise RuntimeError("Chrome or Chromium is required to render the quick-guide PDFs.")

    destination.unlink(missing_ok=True)
    with tempfile.TemporaryDirectory(dir=os.environ.get("XDG_RUNTIME_DIR")) as profile:
        subprocess.run(
            [
                browser,
                "--headless=new",
                "--disable-gpu",
                "--no-first-run",
                "--no-default-browser-check",
                "--no-pdf-header-footer",
                f"--user-data-dir={profile}",
                f"--print-to-pdf={destination}",
                source.resolve().as_uri(),
            ],
            check=True,
            env={**os.environ, "TMPDIR": os.environ.get("XDG_RUNTIME_DIR", os.environ.get("TMPDIR", "/tmp"))},
        )
    if not destination.is_file():
        raise RuntimeError(f"Quick-guide PDF was not created: {destination}")


def build_german_bits_to_words() -> None:
    source = RESOURCES / "BitsToWords.xlsx"
    destination = RESOURCES / "BitsToWords-de.xlsx"
    workbook = load_workbook(source)
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value == "Identifier:":
                    cell.value = "Bezeichnung:"
                elif cell.value == "checksum":
                    cell.value = "Prüfsumme"
    workbook.save(destination)
    convert_to_pdf(destination, ROOT / "BitsToWords-de.pdf")


def build_worksheet_pdfs() -> None:
    convert_to_pdf(RESOURCES / "BitsToWords.xlsx", ROOT / "BitsToWords.pdf")
    build_german_bits_to_words()


def quick_guide_catalog() -> dict[str, object]:
    try:
        catalog = json.loads(QUICK_GUIDE_CATALOG.read_text(encoding="utf-8"))
    except json.JSONDecodeError as error:
        raise RuntimeError(f"Quick-guide catalog is not valid JSON: {QUICK_GUIDE_CATALOG}") from error
    if catalog.get("format") != 1 or not isinstance(catalog.get("html"), dict) or not isinstance(catalog.get("plainText"), dict):
        raise RuntimeError("Quick-guide catalog has an unsupported structure.")
    return catalog


def localized(value: object, locale_code: str) -> str:
    if not isinstance(value, dict) or set(value) != {"en", "de"} or not isinstance(value[locale_code], str):
        raise RuntimeError("Quick-guide catalog translation values must contain paired en/de strings.")
    return value[locale_code]


def validate_quick_guide_catalog(catalog: dict[str, object]) -> None:
    html_content = catalog["html"]
    plain_text = catalog["plainText"]
    if not isinstance(html_content, dict) or not isinstance(plain_text, dict):
        raise RuntimeError("Quick-guide catalog is missing its HTML or plain-text content.")
    for field in ("description", "title"):
        localized(html_content.get(field), "en")

    body = html_content.get("body")
    if not isinstance(body, list) or [entry.get("slot") for entry in body if isinstance(entry, dict)] != list(QUICK_GUIDE_SLOTS):
        raise RuntimeError("Quick-guide HTML sections do not match the shared template plan.")
    for entry in body:
        if not isinstance(entry, dict):
            raise RuntimeError("Quick-guide HTML section is invalid.")
        localized(entry.get("content"), "en")

    localized(plain_text.get("heading"), "en")
    localized(plain_text.get("headingSuffix"), "en")
    sections = plain_text.get("sections")
    if not isinstance(sections, list) or [entry.get("number") for entry in sections if isinstance(entry, dict)] != list(range(1, 9)):
        raise RuntimeError("Quick-guide plain-text sections must be numbered 1 through 8.")
    for entry in sections:
        if not isinstance(entry, dict):
            raise RuntimeError("Quick-guide plain-text section is invalid.")
        localized(entry.get("content"), "en")
        localized(entry.get("suffix"), "en")


def render_quick_guide_html(locale_code: str) -> str:
    catalog = quick_guide_catalog()
    validate_quick_guide_catalog(catalog)
    html_content = catalog["html"]
    assert isinstance(html_content, dict)
    body = html_content["body"]
    assert isinstance(body, list)
    rendered_body = "\n  ".join(localized(entry["content"], locale_code) for entry in body if isinstance(entry, dict))
    template = QUICK_GUIDE_TEMPLATE.read_text(encoding="utf-8")
    return (
        template
        .replace("{{lang}}", locale_code)
        .replace("{{description}}", html.escape(localized(html_content["description"], locale_code), quote=True))
        .replace("{{title}}", html.escape(localized(html_content["title"], locale_code)))
        .replace("{{body}}", f"  {rendered_body}")
    )


def render_quick_guide_text(locale_code: str) -> str:
    catalog = quick_guide_catalog()
    validate_quick_guide_catalog(catalog)
    plain_text = catalog["plainText"]
    assert isinstance(plain_text, dict)
    sections = plain_text["sections"]
    assert isinstance(sections, list)
    output = localized(plain_text["heading"], locale_code) + localized(plain_text["headingSuffix"], locale_code)
    for entry in sections:
        if isinstance(entry, dict):
            output += f'{entry["number"]}. {localized(entry["content"], locale_code)}{localized(entry["suffix"], locale_code)}'
    return output


def quick_guide_name(locale_code: str, suffix: str) -> str:
    locale_suffix = "-de" if locale_code == "de" else ""
    return f"HowToRollYourOwnSeedphrase{locale_suffix}.{suffix}"


def build_quick_guides() -> None:
    for locale_code in ("en", "de"):
        (ROOT / quick_guide_name(locale_code, "html")).write_text(
            render_quick_guide_html(locale_code), encoding="utf-8", newline="\n"
        )
        (ROOT / quick_guide_name(locale_code, "txt")).write_text(
            render_quick_guide_text(locale_code), encoding="utf-8", newline="\n"
        )
        render_html_to_pdf(ROOT / quick_guide_name(locale_code, "html"), ROOT / quick_guide_name(locale_code, "pdf"))


def remove_duplicate_lookup_artifacts() -> None:
    for path in (
        ROOT / "BIP39_Wordlist_Binary_Decimal_Searchable-de.pdf",
        RESOURCES / "BIP39_Wordlist_Binary_Decimal_Searchable-de.ods",
    ):
        path.unlink(missing_ok=True)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("target", nargs="?", choices=("all",), default="all")
    parser.parse_args()
    build_worksheet_pdfs()
    build_quick_guides()
    remove_duplicate_lookup_artifacts()


if __name__ == "__main__":
    main()